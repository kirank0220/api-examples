#########################################################################
#    _________        ___.                   ______________________  ___
#    \_   ___ \___.__.\_ |__   ___________  /  _____/\______   \   \/  /
#    /    \  \<   |  | | __ \_/ __ \_  __ \/   \  ___ |       _/\     /
#    \     \___\___  | | \_\ \  ___/|  | \/\    \_\  \|    |   \/     \
#     \______  / ____| |___  /\___  >__|    \______  /|____|_  /___/\  \
#            \/\/          \/     \/               \/        \/      \_/
#
#

import os
import json
import requests
from datetime import datetime
from collections import OrderedDict
from email_validator import validate_email, EmailNotValidError
from openpyxl import Workbook
from openpyxl.styles.fills import FILL_SOLID
from openpyxl.styles import Color, PatternFill, Font, Border, Side
from openpyxl.styles import colors
from openpyxl.cell import Cell
from tqdm import tqdm
from glom import glom, OMIT, GlomError


VALID_ANSWERS = {
    "least": "Least",
    "minimal": "Minimal",
    "moderate": "Moderate",
    "significant": "Significant",
}

VALID_ASSESSMENT_TIERS = {
    "tier 1": {"tier": 1, "validated": True},  # Tier 1 is always validated
    "tier 2": {"tier": 2, "validated": False},
    "tier 2 validated": {"tier": 2, "validated": True},
    "tier 3": {"tier": 3, "validated": False},  # Tier 3 is never validated
    "no assessment": None,
}

INHERENT_RISK_FROM_RECOMMENDATION = {
    0: "Unconfirmed Exposure",
    1: "High Exposure",
    2: "Medium Exposure",
    3: "Low Exposure",
}


def skip_falsy(value):
    return OMIT if not value else value


def as_string(value):
    return value if not value else str(value)


def insert_http(value):
    if value and not value.startswith("http"):
        return "https://" + value

    return value


def row_to_vendor(row, column_mapping):
    vendor = {column_mapping[cell.column_id]: cell.value for cell in row.cells}
    vendor["custom_id"] = str(row.id)
    return vendor


def normalize_vendor(row, column_mapping, spec):
    vendor = glom(row_to_vendor(row, column_mapping), spec)

    # Check that vendor record for accuracy
    if "url" not in vendor:
        vendor["record_has_url_and_address"] = False
    elif "address" not in vendor:
        vendor["record_has_url_and_address"] = False
    elif "city" not in vendor["address"]:
        vendor["record_has_url_and_address"] = False
    elif "country" not in vendor["address"]:
        vendor["record_has_url_and_address"] = False
    else:
        vendor["record_has_url_and_address"] = True

    return vendor


def lookup_sheet_id(smart, sheet_name):
    response = smart.Sheets.list_sheets(include_all=True)
    matched_sheets = [sheet for sheet in response.data if sheet.name.lower() == sheet_name.lower()]
    if len(matched_sheets) != 1:
        message = (
            "Unable to lookup a unique sheet ID, multiple sheets matched '" + sheet_name + "' set --sheet-id instead"
        )
        raise Exception(message)

    return matched_sheets[0].id


def required(value):
    if not value:
        raise GlomError("Value was not defined")

    return value


def split(last, spliton=" "):
    def parser(value):
        if not value:
            raise GlomError("Value was not defined")

        split = value.split(spliton)
        if len(split) < 2:
            raise GlomError("Value did not have 2 or more elements")

        return split[-1] if last else split[0]

    return parser


def email_metadata(selector):
    first_name = split(False, ".")
    last_name = split(True, ".")

    def parser(value):
        if not value:
            raise GlomError("Value was not defined")

        try:
            parsed = validate_email(value, check_deliverability=False, allow_empty_local=True)

            if selector == "domain":
                return parsed["domain"]

            return (
                first_name(parsed["local"]) if selector == "first_name" else last_name(parsed["local"])
            ).capitalize()
        except EmailNotValidError as e:
            raise GlomError(str(e))

    return parser


def validate_answer(value):
    if not value:
        return OMIT

    normalized = str(value).strip().lower()
    for key in VALID_ANSWERS.keys():
        if normalized.startswith(key):
            return VALID_ANSWERS[key]

    return OMIT


def inherent_risk_level_from_tier(value):
    try:
        return INHERENT_RISK_FROM_RECOMMENDATION[value]
    except KeyError:
        return INHERENT_RISK_FROM_RECOMMENDATION[0]


def valid_assessment_order(value):
    if not value:
        return OMIT

    normalized = str(value).strip().lower()
    try:
        return VALID_ASSESSMENT_TIERS[normalized]
    except KeyError:
        return OMIT


def date_or_none(value):
    return datetime.strptime(value, "%Y-%m-%d") if value else None


def category_match(category):
    def matcher(value):
        if not value:
            return None

        for v in value:
            if v["category"] == category:
                return v

        return None

    return matcher


def _cell_value(cell):
    return "{}".format(cell.value).strip() if cell and cell.value else ""


def sheet_writer(wb, name, columns, mapping=None):
    if not mapping:
        mapping = {}

    for c in columns:
        if not mapping.get(c[1], None):
            mapping[c[1]] = c[1]

    def builder(sheet):
        for idx, injector in enumerate(columns):
            cell = sheet.cell(row=1, column=1 + idx)
            cell.value = injector[0]
            cell.font = cell.font.copy(bold=True)

            if len(injector) <= 2:
                cell.fill = PatternFill(FILL_SOLID, start_color="C9C9C9", end_color="C9C9C9")
            elif injector[2] == "blue":
                cell.fill = PatternFill(FILL_SOLID, start_color="0065B8", end_color="0065B8")
                cell.font = cell.font.copy(color=colors.WHITE)
            elif injector[2] == "orange":
                cell.fill = PatternFill(FILL_SOLID, start_color="FFB802", end_color="FFB802")
            else:
                cell.fill = PatternFill(FILL_SOLID, start_color="C9C9C9", end_color="C9C9C9")

        def write_value(_row, _col, _val):
            cell = sheet.cell(row=_row, column=_col)
            cell.value = _val

        __non_local = {"row": 2}

        def writer(blob):
            transformed = glom(blob, mapping)
            multi_row = 0
            for idx, injector in enumerate(columns):
                value = transformed[injector[1]]
                if value is None:
                    continue

                if not isinstance(value, (list, tuple)):
                    write_value(__non_local["row"], 1 + idx, value)
                else:
                    multi_row = max(multi_row, len(value))
                    for i, v in enumerate(value):
                        write_value(__non_local["row"] + i, 1 + idx, v)

            __non_local["row"] = __non_local["row"] + (multi_row if multi_row else 1)

        def finalizer():
            for column_cells in sheet.columns:
                length = min(125, max(9, max(len(_cell_value(cell)) + 1 for cell in column_cells)))

                for cell in column_cells:
                    cell.alignment = cell.alignment.copy(wrapText=True)

                sheet.column_dimensions[column_cells[0].column_letter].width = length

        writer.finalizer = finalizer
        return writer

    return builder(wb[name])
